#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test CV Data Generator

Generates test CV data in JSON format based on an Excel data matrix.
Reads skill requirements from Excel and generates candidate profiles.
Updates Excel with generation results.
"""

import openpyxl
import json
from datetime import datetime
from pathlib import Path


class CVTestDataGenerator:
    """Generate test CV data from Excel skill matrix"""

    # Static candidate data
    STATIC_CV_ID = "12788"
    STATIC_NAME = "平林 直美"
    STATIC_FURIGANA = "ひらばやし なおみ"
    STATIC_BIRTHDATE = "1977-07-20"
    STATIC_GENDER = "女"
    STATIC_ADDRESS = "東京都国分寺市西恋ヶ窪2-8-11"

    # Static JD (Job Description) data
    STATIC_JD_CONTENT = (
        "美術科 非常勤講師（中高一貫校）／勤務地: 東京都世田谷区／コマ数: 週13コマ／開始: 11月から\n"
        "必須: 中学・高校の美術教員免許。\n"
        "補足: ICT導入校。"
    )

    # Static weights for matching algorithm
    STATIC_TOP_K = 1
    STATIC_CUSTOM_WEIGHTS = {
        "license_score": 0.4,
        "hours_score": 0.2,
        "reliability_score": 0.2,
        "competency_score": 0.2
    }

    RESULT_COLUMN_HEADER = "cv generated result"

    def __init__(self, excel_path: str):
        """
        Initialize generator with Excel file path

        Args:
            excel_path: Path to Excel file containing skill matrix
        """
        self.excel_path = Path(excel_path)
        self.workbook = None
        self.worksheet = None
        self.category_headers = {}  # Row 1: merged category headers
        self.skill_headers = []      # Row 2: detailed skill headers
        self.university_details = {}
        self.result_column_idx = None

    def load_excel(self):
        """Load Excel workbook and extract headers"""
        self.workbook = openpyxl.load_workbook(self.excel_path)
        self.worksheet = self.workbook.active

        # Extract category headers from row 1 (merged cells)
        # Columns A-E (1-5): "資格・免許"
        # Columns F-I (6-9): "大学名"
        self.category_headers = {
            'certificate': {
                'name': '資格・免許',
                'start_col': 1,
                'end_col': 5
            },
            'university': {
                'name': '大学名',
                'start_col': 6,
                'end_col': 9
            }
        }

        # Extract detailed skill headers from row 2
        for col_idx in range(1, self.worksheet.max_column + 1):
            header_value = self.worksheet.cell(2, col_idx).value
            # Strip whitespace from headers
            clean_header = header_value.strip() if header_value else ""
            self.skill_headers.append(clean_header)

        # Check if result column exists, if not add it
        self._ensure_result_column()

    def _ensure_result_column(self):
        """Add result column if it doesn't exist"""
        for idx, header in enumerate(self.skill_headers):
            if header and self.RESULT_COLUMN_HEADER in str(header).lower():
                self.result_column_idx = idx + 1
                return

        new_col_idx = len(self.skill_headers) + 1
        self.result_column_idx = new_col_idx
        self.worksheet.cell(2, new_col_idx).value = self.RESULT_COLUMN_HEADER
        self.skill_headers.append(self.RESULT_COLUMN_HEADER)

    def _update_result_column(self, row_idx: int, cv_id: str, cv_content: str):
        """
        Update result column for generated CV with complete JSON structure

        Args:
            row_idx: Excel row index (1-based)
            cv_id: Generated CV ID
            cv_content: Full CV content string
        """
        # Create the complete JSON structure for this single CV
        result_json = {
            "list_cv": [
                {
                    "cv_id": cv_id,
                    "content": cv_content
                }
            ],
            "jd": {
                "content": self.STATIC_JD_CONTENT
            },
            "top_k": self.STATIC_TOP_K,
            "custom_weights": self.STATIC_CUSTOM_WEIGHTS
        }

        # Convert to JSON string with proper formatting
        result_text = json.dumps(result_json, ensure_ascii=False, indent=2)
        self.worksheet.cell(row_idx, self.result_column_idx).value = result_text

    def save_excel(self):
        """Save changes to Excel file"""
        self.workbook.save(self.excel_path)

    def extract_candidate_skills(self, row_idx: int) -> dict:
        """
        Extract skills for a candidate from Excel row

        Args:
            row_idx: Row index in Excel (1-based, starting from row 4 for candidate data)

        Returns:
            Dictionary with 'certificate' and 'university' lists of skill names
        """
        certificate_skills = []
        university_skills = []

        # Extract certificate skills (columns A-E, 1-5)
        for col_idx in range(1, 6):
            if col_idx == self.result_column_idx:
                continue
            cell_value = self.worksheet.cell(row_idx, col_idx).value
            if str(cell_value).strip().lower() == "x":
                skill_name = self.skill_headers[col_idx - 1]
                if skill_name:
                    certificate_skills.append(skill_name)

        # Extract university skills (columns F-I, 6-9)
        for col_idx in range(6, 10):
            if col_idx == self.result_column_idx:
                continue
            if col_idx > len(self.skill_headers):
                break
            cell_value = self.worksheet.cell(row_idx, col_idx).value
            if str(cell_value).strip().lower() == "x":
                skill_name = self.skill_headers[col_idx - 1]
                if skill_name:
                    university_skills.append(skill_name)

        return {
            'certificate': certificate_skills,
            'university': university_skills
        }

    def build_cv_content(self, skills: dict) -> str:
        """
        Build CV content string based on skills

        Args:
            skills: Dictionary with 'certificate' and 'university' skill lists

        Returns:
            Formatted CV content string with grouped skills
        """
        content = f"氏名: {self.STATIC_NAME}\n"
        content += f"フリガナ: {self.STATIC_FURIGANA}\n"
        content += f"性別: {self.STATIC_GENDER}\n"
        content += f"生年月日: {self.STATIC_BIRTHDATE}\n"
        content += f"住所: {self.STATIC_ADDRESS}\n"

        # Add certificate skills as a comma-separated group
        if skills['certificate']:
            certificate_str = ', '.join(skills['certificate'])
            content += f"資格・免許: {certificate_str}\n"

        # Add university skills as a comma-separated group
        if skills['university']:
            university_str = ', '.join(skills['university'])
            content += f"大学名: {university_str}\n"

        return content

    def generate_single_cv(self, row_idx: int, update_excel: bool = True) -> dict:
        """
        Generate single CV data from Excel row

        Args:
            row_idx: Row index in Excel (1-based)
            update_excel: Whether to update Excel with generation result

        Returns:
            Dictionary containing CV data
        """
        cv_id = self.STATIC_CV_ID
        skills = self.extract_candidate_skills(row_idx)
        content = self.build_cv_content(skills)

        if update_excel:
            self._update_result_column(row_idx, cv_id, content)

        return {
            "cv_id": cv_id,
            "content": content
        }

    def generate_all_cvs(self, output_path: str = None, update_excel: bool = True) -> dict:
        """
        Generate CV data for all candidates in Excel

        Args:
            output_path: Optional path to save JSON output
            update_excel: Whether to update Excel with generation results

        Returns:
            Dictionary containing list_cv, jd, top_k, and custom_weights
        """
        self.load_excel()

        cvs = []
        # Start from row 3 (row 1: category headers, row 2: skill headers, row 3+: candidates)
        for row_idx in range(3, self.worksheet.max_row + 1):
            has_data = any(
                self.worksheet.cell(row_idx, col).value
                for col in range(1, self.worksheet.max_column + 1)
                if col != self.result_column_idx
            )

            if has_data:
                cv_data = self.generate_single_cv(row_idx, update_excel)
                cvs.append(cv_data)

        if update_excel:
            self.save_excel()

        # Create the complete output structure
        output_data = {
            "list_cv": cvs,
            "jd": {
                "content": self.STATIC_JD_CONTENT
            },
            "top_k": self.STATIC_TOP_K,
            "custom_weights": self.STATIC_CUSTOM_WEIGHTS
        }

        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, ensure_ascii=False, indent=2)

        return output_data

    def generate_cv_by_index(self, candidate_index: int, update_excel: bool = True) -> dict:
        """
        Generate CV for specific candidate by index

        Args:
            candidate_index: Candidate index (0-based)
            update_excel: Whether to update Excel with generation result

        Returns:
            CV dictionary
        """
        self.load_excel()
        row_idx = candidate_index + 3  # Row 3 is first candidate (0-based index)

        if row_idx > self.worksheet.max_row:
            raise IndexError(f"Candidate index {candidate_index} out of range")

        cv_data = self.generate_single_cv(row_idx, update_excel)

        if update_excel:
            self.save_excel()

        return cv_data


def main():
    """Main entry point for command-line usage"""
    import argparse

    parser = argparse.ArgumentParser(description='Generate test CV data from Excel skill matrix')
    parser.add_argument('excel_file', help='Path to Excel file')
    parser.add_argument('-o', '--output', help='Output JSON file path', default='cv_test_data.json')
    parser.add_argument('-n', '--number', type=int, help='Generate specific candidate by index (0-based)')
    parser.add_argument('--no-update', action='store_true', help='Do not update Excel with generation results')

    args = parser.parse_args()

    generator = CVTestDataGenerator(args.excel_file)
    update_excel = not args.no_update

    if args.number is not None:
        cv_data = generator.generate_cv_by_index(args.number, update_excel)
        print(json.dumps(cv_data, ensure_ascii=False, indent=2))
        if update_excel:
            print(f"\nUpdated Excel file: {args.excel_file}")
    else:
        output_data = generator.generate_all_cvs(args.output, update_excel)
        print(f"Generated {len(output_data['list_cv'])} CV records")
        print(f"Saved to: {args.output}")
        if update_excel:
            print(f"Updated Excel file: {args.excel_file}")

        if output_data['list_cv']:
            print("\nExample (first CV):")
            print(json.dumps(output_data['list_cv'][0], ensure_ascii=False, indent=2))
            print("\nComplete structure preview:")
            print(f"- list_cv: {len(output_data['list_cv'])} CVs")
            print(f"- jd: Job description included")
            print(f"- top_k: {output_data['top_k']}")
            print(f"- custom_weights: {output_data['custom_weights']}")


if __name__ == '__main__':
    main()
